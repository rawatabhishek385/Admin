# admin.py
from __future__ import annotations
from django.contrib import admin, messages
from django.db import transaction
from django.shortcuts import render, redirect
from django.urls import path, reverse
from django.http import HttpResponse, HttpResponseForbidden
from django.template.response import TemplateResponse
import time

from .models import Candidate, Question, Answer
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side


# ------------ Excel helpers ------------

REQUIRED_COLS = {"army_no", "exam_type", "question", "answer"}
KNOWN_COLS = {
    "s_no", "name","center" ,"photo", "fathers_name", "dob", "trade","rank", "army_no", "adhaar_no",  # Added trade
    "name_of_qualification", "duration_of_qualification", "credits", "nsqf_level",
    "training_center", "district", "state", "viva_1", "viva_2",
    "practical_1", "practical_2", "exam_type", "question",
    "answer", "correct_answer", "max_marks",
}


def _normalize_header(val: str) -> str:
    return (
        (val or "")
        .strip()
        .lower()
        .replace(".", "_")
        .replace(" ", "_")
    )


def _read_rows_from_excel(file):
    wb = load_workbook(file, data_only=True)
    ws = wb.worksheets[0]

    headers = [_normalize_header(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
    header_index = {h: idx for idx, h in enumerate(headers) if h}

    missing = REQUIRED_COLS - set(header_index)
    if missing:
        raise ValueError(f"Missing required columns in Excel: {', '.join(missing)}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {}
        for key, idx in header_index.items():
            data[key] = row[idx]
        yield data


def _get_or_create_question(exam_type, text, correct, max_marks):
    q = Question.objects.filter(exam_type=exam_type, question=text).first()
    # normalize correct_answer
    correct_clean = (correct or "")
    if isinstance(correct_clean, str) and correct_clean.strip().lower() == "null":
        correct_clean = None

    if q is None:
        q = Question.objects.create(
            exam_type=exam_type,
            question=text,
            correct_answer=correct_clean,
            max_marks=max_marks or 0,
        )
    else:
        # update existing question too
        q.correct_answer = correct_clean
        q.max_marks = max_marks or 0
        q.save()
    return q


# ------------ Custom Admins ------------

class AnswerInline(admin.TabularInline):
    model = Answer
    extra = 0


@admin.register(Candidate)
class CandidateAdmin(admin.ModelAdmin):
    change_list_template = "admin/exams/candidate/change_list.html"
    change_form_template = "admin/exams/candidate/change_form.html"
    list_display = ("army_no", "name","center", "trade", "total_primary", "total_secondary", "grand_total")  # Added trade
    list_filter = ("center","trade")  # Added trade filter
    search_fields = ("army_no", "name","rank", "fathers_name", "district", "state", "trade")  # Added trade

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path("import-excel/", self.admin_site.admin_view(self.import_excel_view),
                 name="exams_candidate_import_excel"),
            path("export-results-excel/", self.admin_site.admin_view(self.export_results_excel_view),
                 name="exams_export_results_excel"),
            path("<int:candidate_id>/save-grades/", self.admin_site.admin_view(self.save_grades_view),
                 name="exams_candidate_save_grades"),
            path("<int:candidate_id>/grade-answers/", self.admin_site.admin_view(self.grade_answers_view),
                 name="exams_candidate_grade_answers"),
        ]
        return custom + urls

    # ---------- Candidate change form ----------
    def change_view(self, request, object_id, form_url="", extra_context=None):
        cand = Candidate.objects.get(pk=object_id)
        answers = Answer.objects.filter(candidate=cand).select_related("question")

        primary = [a for a in answers if a.question.exam_type.lower() == "primary"]
        secondary = [a for a in answers if a.question.exam_type.lower() == "secondary"]

        extra_context = extra_context or {}
        extra_context["primary_answers"] = primary
        extra_context["secondary_answers"] = secondary
        extra_context["viva_total"] = cand.viva_1 + cand.viva_2
        extra_context["practical_total"] = cand.practical_1 + cand.practical_2
        extra_context["show_grade_button"] = True
        
        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    # ---------- Grade Answers View ----------
    def grade_answers_view(self, request, candidate_id):
        if not request.user.has_perm('exams.change_answer'):
            return HttpResponseForbidden("You don't have permission to grade answers")
        
        cand = Candidate.objects.get(pk=candidate_id)
        answers = Answer.objects.filter(candidate=cand).select_related("question")
        
        # 🔥 Auto-marking logic
        for ans in answers:
            cand_ans = (ans.answer or "").strip().lower()
            corr_raw = (ans.question.correct_answer or "").strip().lower()
            if cand_ans and corr_raw:
                correct_list = [c.strip() for c in corr_raw.split(",")]
                if cand_ans in correct_list:
                    if ans.marks_obt is None or ans.marks_obt == 0:
                        ans.marks_obt = ans.question.max_marks
                        ans.save()
        
        # Separate answers by exam type
        primary_answers = [a for a in answers if a.question.exam_type.lower() == "primary"]
        secondary_answers = [a for a in answers if a.question.exam_type.lower() == "secondary"]
        
        if request.method == "POST":
            # Process the form submission
            for answer in answers:
                field_name = f"marks_{answer.id}"
                if field_name in request.POST:
                    try:
                        new_marks = int(request.POST[field_name])
                        if 0 <= new_marks <= answer.question.max_marks:
                            answer.marks_obt = new_marks
                            answer.save()
                    except ValueError:
                        pass
            
            # Force refresh of the candidate object to ensure latest data
            cand = Candidate.objects.get(pk=candidate_id)
            
            self.message_user(request, "Grades updated successfully", level=messages.SUCCESS)
            # Add a cache-busting parameter to the redirect URL
            return redirect(f"{reverse('admin:exams_candidate_change', args=[candidate_id])}?t={time.time()}")
        
        primary_total_obtained = sum(a.marks_obt or 0 for a in primary_answers)
        primary_total_max = sum(a.question.max_marks for a in primary_answers)
        secondary_total_obtained = sum(a.marks_obt or 0 for a in secondary_answers)
        secondary_total_max = sum(a.question.max_marks for a in secondary_answers)
        
        context = {
            **self.admin_site.each_context(request),
            "title": f"Grade Answers - {cand.name} ({cand.army_no}) - {cand.trade}",  # Added trade
            "candidate": cand,
            "primary_answers": primary_answers,
            "secondary_answers": secondary_answers,
            "primary_total_obtained": primary_total_obtained,
            "primary_total_max": primary_total_max,
            "secondary_total_obtained": secondary_total_obtained,
            "secondary_total_max": secondary_total_max,
            "opts": self.model._meta,
        }
        
        return TemplateResponse(request, "admin/exams/candidate/grade_answers.html", context)

    def save_grades_view(self, request, candidate_id):
        cand = Candidate.objects.get(pk=candidate_id)
        if request.method == "POST":
            for ans in Answer.objects.filter(candidate=cand):
                field_name = f"marks_{ans.id}"
                if field_name in request.POST:
                    try:
                        new_marks = int(request.POST[field_name])
                        ans.marks_obt = new_marks
                        ans.save()
                    except ValueError:
                        pass
            self.message_user(request, "Grades updated", level=messages.SUCCESS)
        return redirect("admin:exams_candidate_change", cand.id)

    # ---------- Import Excel ----------
    def import_excel_view(self, request):
        if request.method == "POST" and request.FILES.get("excel"):
            excel_file = request.FILES["excel"]
            created_candidates = updated_candidates = 0
            created_answers = updated_answers = 0
            created_questions = 0

            try:
                with transaction.atomic():
                    seen_questions_before = set(Question.objects.values_list("id", flat=True))
                    for row in _read_rows_from_excel(excel_file):
                        army = (row.get("army_no") or "").strip()
                        if not army:
                            continue

                        cand_defaults = {
                            "s_no": row.get("s_no") or 0,
                            "name": row.get("name") or "",
                            "center": row.get("center") or "",  # Added center
                            "photo": row.get("photo") or None,
                            "fathers_name": row.get("fathers_name") or "",
                            "dob": row.get("dob") or None,
                            "rank": row.get("rank") or "",  # Added rank
                            "trade": row.get("trade") or "",  # Added trade
                            "adhaar_no": row.get("adhaar_no") or "",
                            "name_of_qualification": row.get("name_of_qualification") or "",
                            "duration_of_qualification": row.get("duration_of_qualification") or "",
                            "credits": row.get("credits") or 0,
                            "nsqf_level": row.get("nsqf_level") or 0,
                            "training_center": row.get("training_center") or "",
                            "district": row.get("district") or "",
                            "state": row.get("state") or "",
                            "viva_1": row.get("viva_1") or 0,
                            "viva_2": row.get("viva_2") or 0,
                            "practical_1": row.get("practical_1") or 0,
                            "practical_2": row.get("practical_2") or 0,
                        }
                        cand, created = Candidate.objects.get_or_create(
                            army_no=army, defaults=cand_defaults
                        )
                        if not created:
                            for k, v in cand_defaults.items():
                                if v and getattr(cand, k) != v:
                                    setattr(cand, k, v)
                            cand.save()
                            updated_candidates += 1
                        else:
                            created_candidates += 1

                        q = _get_or_create_question(
                            exam_type=row.get("exam_type") or "",
                            text=row.get("question") or "",
                            correct=row.get("correct_answer") ,
                            max_marks=row.get("max_marks") or 0,
                        )
                        if q.id not in seen_questions_before:
                            created_questions += 1
                            seen_questions_before.add(q.id)

                        ans_text = (row.get("answer") or "").strip()
                        marks = row.get("marks_obt") or 0

                        ans, a_created = Answer.objects.get_or_create(
                            candidate=cand, question=q,
                            defaults={"answer": ans_text, "marks_obt": int(marks)},
                        )
                        if a_created:
                            created_answers += 1
                        else:
                            if ans.answer != ans_text or ans.marks_obt != int(marks):
                                ans.answer = ans_text
                                ans.marks_obt = int(marks)
                                ans.save()
                                updated_answers += 1

                self.message_user(
                    request,
                    (
                        f"Import complete. "
                        f"Candidates: +{created_candidates} / updated {updated_candidates}. "
                        f"Questions: +{created_questions}. "
                        f"Answers: +{created_answers} / updated {updated_answers}."
                    ),
                    level=messages.SUCCESS,
                )
                return redirect("admin:exams_candidate_changelist")

            except Exception as e:
                self.message_user(request, f"Import failed: {e}", level=messages.ERROR)

        ctx = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Import candidates & answers from Excel",
        }
        return render(request, "admin/exams/candidate/import_excel.html", ctx)

    # ---------- Export ----------


    def export_results_excel_view(self, request):
        wb = Workbook()
        
        # Create Primary sheet
        ws_primary = wb.active
        ws_primary.title = "PRIMARY MARKS STATEMENT"
        
        # Create Secondary sheet
        ws_secondary = wb.create_sheet(title="SECONDARY MARKS STATEMENT")

        # Create Combined sheet
        # --- Combined Sheet ---
        ws_combined = wb.create_sheet(title="COMBINED RESULTS")

# ---------- Styles ----------
        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ----- Combined Sheet Formatting -----
        ws_combined.merge_cells("A1:A2")  # S No
        ws_combined.merge_cells("B1:B2")  # Centre
        ws_combined.merge_cells("C1:C2")  # Army No
        ws_combined.merge_cells("D1:D2")  # Rk
        ws_combined.merge_cells("E1:E2")  # Tde
        ws_combined.merge_cells("F1:F2")  # Name

        # Get qualification names dynamically
        first_cand = Candidate.objects.first()
        primary_qf_name = f"Primary-1" if first_cand else "Primary-1"
        secondary_qf_name = f"Secondary-1" if first_cand else "Secondary-1"

        # Merge headers for primary and secondary blocks (5 cols each)
        ws_combined.merge_cells("G1:K1")  # Primary-1
        ws_combined.merge_cells("L1:P1")  # Secondary-1

        # --------- Write Row 1 (main headers) ---------
        ws_combined["A1"] = "S No"
        ws_combined["B1"] = "Centre"
        ws_combined["C1"] = "Army No"
        ws_combined["D1"] = "Rk"
        ws_combined["E1"] = "Tde"
        ws_combined["F1"] = "Name"
        ws_combined["G1"] = primary_qf_name
        ws_combined["L1"] = secondary_qf_name

        # --------- Write Row 2 (sub-headers) ---------
        sub_headers = [
            "Theory*", "Practical*", "Viva*", "Total", "Percentage (%)",
            "Theory*", "Practical*", "Viva*", "Total", "Percentage (%)"
        ]
        for i, val in enumerate(sub_headers, start=7):  # G=7th col
            ws_combined.cell(row=2, column=i, value=val)

        # Apply header styles
        for row in ws_combined.iter_rows(min_row=1, max_row=2):
            for cell in row:
                if cell.value:
                    cell.font = bold_font
                    cell.alignment = center_aligned
                    cell.border = thin_border

        # -------- Fill candidate data --------
        row_idx = 3
        for cand in Candidate.objects.all():
            # Primary
            primary_theory = sum(
                a.marks_obt or 0 for a in cand.answer_set.filter(question__exam_type__iexact="primary")
            )
            primary_practical = cand.practical_1 or 0
            primary_viva = cand.viva_1 or 0
            primary_total = primary_theory + primary_practical + primary_viva
            primary_percentage = primary_total  # adjust denominator if needed

            # Secondary
            secondary_theory = sum(
                a.marks_obt or 0 for a in cand.answer_set.filter(question__exam_type__iexact="secondary")
            )
            secondary_practical = cand.practical_2 or 0
            secondary_viva = cand.viva_2 or 0
            secondary_total = secondary_theory + secondary_practical + secondary_viva
            secondary_percentage = secondary_total  # adjust denominator if needed

            ws_combined.append([
                cand.s_no or "",
                cand.center or "",
                cand.army_no or "",
                cand.rank or "",
                cand.trade or "",
                cand.name or "",
                primary_theory, primary_practical, primary_viva,
                primary_total, primary_percentage,
                secondary_theory, secondary_practical, secondary_viva,
                secondary_total, secondary_percentage
            ])
            row_idx += 1

        # Apply borders to all data rows
        for row in ws_combined.iter_rows(min_row=1, max_row=ws_combined.max_row):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border


        # ---------- Primary & Secondary Headers ----------
        primary_headers = [
            "S No", "Name of Candidate", "Photograph", "Father's Name","Trade", "DOB", 
            "Enrolment No", "Aadhar Number", "Name of Qualification", 
            "Duration of Qualification", "Credits", "NSQF Level", "Training Centre", 
            "District", "State", "Percentage"
        ]
        
        secondary_headers = [
            "S No", "Name of Candidate", "Photograph", "Father's Name","Trade", "DOB", 
            "Enrolment No", "Aadhar Number", "Name of Qualification", 
            "Duration of Qualification", "Credits", "NSQF Level", "Training Centre", 
            "District", "State", "Percentage"
        ]
        
        # Add headers
        ws_primary.append(primary_headers)
        ws_secondary.append(secondary_headers)
        
        # Apply styling to headers
        for row in ws_primary.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = bold_font
                cell.alignment = center_aligned
                cell.border = thin_border
        
        for row in ws_secondary.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = bold_font
                cell.alignment = center_aligned
                cell.border = thin_border
        
        # Process candidates for both sheets
        for cand in Candidate.objects.all():
            primary_total = sum(
                a.marks_obt or 0 for a in cand.answer_set.filter(question__exam_type__iexact="primary")
            ) + (cand.viva_1 or 0) + (cand.practical_1 or 0)
            
            secondary_total = sum(
                a.marks_obt or 0 for a in cand.answer_set.filter(question__exam_type__iexact="secondary")
            ) + (cand.viva_2 or 0) + (cand.practical_2 or 0)
            
            ws_primary.append([
                cand.s_no or "",
                cand.name or "",
                "",  # Photograph placeholder
                cand.fathers_name or "",
                cand.trade or "",
                cand.dob or "",
                cand.army_no or "",
                cand.adhaar_no or "",
                cand.name_of_qualification or "",
                cand.duration_of_qualification or "",
                cand.credits or 0,
                cand.nsqf_level or 0,
                cand.training_center or "",
                cand.district or "",
                cand.state or "",
                primary_total
            ])
            
            ws_secondary.append([
                cand.s_no or "",
                cand.name or "",
                "",  # Photograph placeholder
                cand.fathers_name or "",
                cand.trade or "",
                cand.dob or "",
                cand.army_no or "",
                cand.adhaar_no or "",
                cand.name_of_qualification or "",
                cand.duration_of_qualification or "",
                cand.credits or 0,
                cand.nsqf_level or 0,
                cand.training_center or "",
                cand.district or "",
                cand.state or "",
                secondary_total
            ])
        
        # Borders
        for row in ws_primary.iter_rows(min_row=1, max_row=ws_primary.max_row):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
        
        for row in ws_secondary.iter_rows(min_row=1, max_row=ws_secondary.max_row):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
        
        # Auto-adjust widths
        for sheet in [ws_primary, ws_secondary]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                sheet.column_dimensions[column_letter].width = max_length + 2

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="marks_statements.xlsx"'
        wb.save(resp)
        return resp

    # ---------- Totals ----------
    def total_primary(self, obj):
        return sum(a.marks_obt or 0 for a in obj.answer_set.filter(question__exam_type__iexact="primary"))

    def total_secondary(self, obj):
        return sum(a.marks_obt or 0 for a in obj.answer_set.filter(question__exam_type__iexact="secondary"))

    def grand_total(self, obj):
        viva_practical = (obj.viva_1 or 0) + (obj.viva_2 or 0) + (obj.practical_1 or 0) + (obj.practical_2 or 0)
        return self.total_primary(obj) + self.total_secondary(obj) + viva_practical


# remove Question from sidebar
try:
    admin.site.unregister(Question)
except admin.sites.NotRegistered:
    pass
